#!/usr/bin/env python3
"""
Build Attira_Growth_Dashboard.xlsx — the one-stop weekly growth dashboard.

You type raw social numbers (likes / comments / shares / saves / impressions) into
the YELLOW input cells; the BLUE cells compute engagement, the funnel, per-channel
performance, and — the headline — your pace toward 2,000 waitlist signups before
the App Store launch.

North Star (current phase): Waitlist signups / week.
Beta / TestFlight activation is Phase 2 and intentionally not tracked here yet.

Sheets:
  1. README           — how to use, colour legend, weekly ritual
  2. Post Log         — one row per social post (manual)
  3. Weekly Inputs    — one row per week; PostHog numbers (manual) + auto roll-ups
  4. Funnel Dashboard — all-time funnel + conversion ratios (auto)
  5. Path to 2,000    — cumulative vs goal, required pace, run-rate, RAG (auto)
  6. Channel Scorecard— per-platform comparison (auto + 1 manual signups column)
  7. UTM Links        — share-link catalog (adds Threads)
  8. Targets          — goal, launch date, KPI targets (manual)

Run:  .venv-growth/bin/python scripts/build_growth_dashboard.py
"""

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, Reference

# ── Brand palette (matches build_utm_workbook.py) ──────────────────────────────
BRAND = "1D4AFF"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color=BRAND)
SUB_FONT = Font(italic=True, size=9, color="666666")
BOLD = Font(bold=True)
SECTION_FONT = Font(bold=True, size=12, color="222222")

INPUT_FILL = PatternFill("solid", fgColor="FFF3C4")   # yellow  = you type here
OUTPUT_FILL = PatternFill("solid", fgColor="EAF0FF")  # blue    = formula / read only
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GREEN_FONT = Font(color="006100", bold=True)
AMBER_FONT = Font(color="9C6500", bold=True)
RED_FONT = Font(color="9C0006", bold=True)

THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
NUM = "#,##0"
DATEFMT = "yyyy-mm-dd"

PLATFORMS = ["Reddit", "Twitter/X", "Threads", "Facebook", "LinkedIn", "Instagram", "Product Hunt"]
# platform label -> utm_source (for the UTM Links sheet / reference)
PLATFORM_SOURCE = {
    "Reddit": "reddit", "Twitter/X": "twitter", "Threads": "threads",
    "Facebook": "facebook", "LinkedIn": "linkedin", "Instagram": "instagram",
    "Product Hunt": "producthunt",
}

POSTLOG_FIRST = 5          # first data row on Post Log
POSTLOG_ROWS = 80          # number of blank post rows
POSTLOG_LAST = POSTLOG_FIRST + POSTLOG_ROWS - 1
WEEKS = 16
WEEK_FIRST = 5             # first data row on Weekly Inputs
WEEK_LAST = WEEK_FIRST + WEEKS - 1


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if sub:
        ws["A2"] = sub
        ws["A2"].font = SUB_FONT


wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — README
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "README"
title(ws, "Attira — Growth Dashboard",
      "One sheet to turn social posts into a clear read on progress toward 2,000 waitlist signups.")
lines = [
    ("", ""),
    ("THE GOAL", "section"),
    ("Reach 2,000 waitlist signups before the official App Store launch.", ""),
    ("North Star metric (this phase): Waitlist signups per week.", ""),
    ("Beta / TestFlight tracking is Phase 2 — it switches on once the app is shared by link.", ""),
    ("", ""),
    ("HOW TO USE IT (3 things, weekly)", "section"),
    ("1.  After each post, add a row to 'Post Log' — likes, comments, shares, saves, impressions.", ""),
    ("2.  Once a week, open PostHog and put visitors + signups for that week into 'Weekly Inputs'.", ""),
    ("3.  Once a week, copy each platform's cumulative signups into the yellow column on 'Channel Scorecard'.", ""),
    ("Then just READ: 'Path to 2,000' and 'Funnel Dashboard' update themselves.", ""),
    ("", ""),
    ("COLOUR LEGEND", "section"),
    ("Yellow cells = you type here (raw data).", "inkey"),
    ("Blue cells = calculated for you. Don't overwrite — they hold formulas.", "outkey"),
    ("", ""),
    ("THE 20-MINUTE MONDAY REVIEW (read in this order)", "section"),
    ("1.  Path to 2,000 -> are we ON PACE or BEHIND? (the single most important cell)", ""),
    ("2.  Weekly Inputs -> signups this week vs last week (growing?).", ""),
    ("3.  Channel Scorecard -> which platform drove the most signups per 1k impressions?", ""),
    ("4.  Double down on the top channel; cut or rework the worst.", ""),
    ("", ""),
    ("WHERE THE NUMBERS COME FROM", "section"),
    ("Likes / comments / shares / saves / impressions  ->  each platform's own post analytics.", ""),
    ("Visitors, link clicks, signups by source  ->  PostHog (filter by utm_source).", ""),
    ("Share the UTM links from the 'UTM Links' tab so PostHog can attribute traffic.", ""),
]
r = 4
for text, kind in lines:
    cell = ws.cell(row=r, column=1, value=text)
    if kind == "section":
        cell.font = SECTION_FONT
    elif kind == "inkey":
        cell.fill = INPUT_FILL
    elif kind == "outkey":
        cell.fill = OUTPUT_FILL
    r += 1
widths(ws, [110])

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Post Log  (manual entry, one row per post)
# ══════════════════════════════════════════════════════════════════════════════
pl = wb.create_sheet("Post Log")
title(pl, "Post Log",
      "One row per social post. Fill the yellow columns; Engagement rate & CTR fill in automatically.")
pl_headers = ["Date", "Platform", "Handle / Subreddit", "Post URL", "Content type",
              "Impressions", "Likes", "Comments", "Shares", "Saves", "Link clicks",
              "utm_source", "utm_content", "Engagement rate", "CTR"]
HROW = 4
for i, h in enumerate(pl_headers, start=1):
    pl.cell(row=HROW, column=i, value=h)
style_header(pl, HROW, len(pl_headers))

# Engagement rate = (Likes+Comments+Shares+Saves)/Impressions ; CTR = Link clicks/Impressions
for row in range(POSTLOG_FIRST, POSTLOG_LAST + 1):
    for col in range(1, 14):          # A..M are input
        c = pl.cell(row=row, column=col)
        c.fill = INPUT_FILL
        c.border = BORDER
    eng = pl.cell(row=row, column=14,
                  value=f'=IFERROR((G{row}+H{row}+I{row}+J{row})/F{row},"")')
    ctr = pl.cell(row=row, column=15, value=f'=IFERROR(K{row}/F{row},"")')
    for c in (eng, ctr):
        c.fill = OUTPUT_FILL
        c.border = BORDER
        c.number_format = PCT
    pl.cell(row=row, column=1).number_format = DATEFMT

# Platform dropdown
dv = DataValidation(type="list", formula1='"%s"' % ",".join(PLATFORMS), allow_blank=True)
pl.add_data_validation(dv)
dv.add(f"B{POSTLOG_FIRST}:B{POSTLOG_LAST}")
pl.freeze_panes = f"A{POSTLOG_FIRST}"
widths(pl, [12, 13, 18, 30, 16, 12, 9, 11, 9, 9, 11, 13, 16, 15, 9])

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 8 (built early so others can reference it) — Targets
# ══════════════════════════════════════════════════════════════════════════════
tg = wb.create_sheet("Targets")
title(tg, "Targets", "Set the goal and your weekly bars here. These drive the RAG colours everywhere else.")
tg["A3"] = "Goal — total waitlist signups";            tg["B3"] = 2000
tg["A4"] = "Signups already captured (baseline)";      tg["B4"] = 8
tg["A5"] = "Target App Store launch date";             tg["B5"] = datetime.date.today() + datetime.timedelta(weeks=8)
tg["A7"] = "Weekly KPI targets"; tg["A7"].font = SECTION_FONT
tg["A8"] = "Engagement rate";                          tg["B8"] = 0.04
tg["A9"] = "Click-through rate (CTR)";                 tg["B9"] = 0.015
tg["A10"] = "Visitor -> Signup conversion";            tg["B10"] = 0.15
for rr in (3, 4, 5, 8, 9, 10):
    tg.cell(row=rr, column=1).font = BOLD
    bcell = tg.cell(row=rr, column=2)
    bcell.fill = INPUT_FILL
    bcell.border = BORDER
tg["B3"].number_format = NUM
tg["B4"].number_format = NUM
tg["B5"].number_format = DATEFMT
for rr in (8, 9, 10):
    tg.cell(row=rr, column=2).number_format = PCT
tg["A12"] = "Yellow = edit these. Launch date defaults to 8 weeks out — change it to your real date."
tg["A12"].font = SUB_FONT
widths(tg, [36, 16])

GOAL = "Targets!$B$3"
BASE = "Targets!$B$4"
LAUNCH = "Targets!$B$5"
T_ENG = "Targets!$B$8"
T_CTR = "Targets!$B$9"
T_CONV = "Targets!$B$10"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Weekly Inputs
# ══════════════════════════════════════════════════════════════════════════════
wi = wb.create_sheet("Weekly Inputs")
title(wi, "Weekly Inputs",
      "Once a week: enter the week's start date + PostHog visitors & signups (yellow). The rest rolls up from Post Log.")
wi_headers = ["Week #", "Week starting", "Impressions", "Engagements", "Link clicks (social)",
              "Qualified visitors", "Waitlist signups", "Cumulative signups",
              "Engagement rate", "Visitor->Signup %"]
HROW = 4
for i, h in enumerate(wi_headers, start=1):
    wi.cell(row=HROW, column=i, value=h)
style_header(wi, HROW, len(wi_headers))

# prefill week numbers + Monday-aligned start dates for convenience
this_monday = datetime.date.today() - datetime.timedelta(days=datetime.date.today().weekday())
PL = "'Post Log'"
for idx in range(WEEKS):
    row = WEEK_FIRST + idx
    wi.cell(row=row, column=1, value=idx + 1).font = BOLD
    dcell = wi.cell(row=row, column=2, value=this_monday + datetime.timedelta(weeks=idx))
    dcell.number_format = DATEFMT
    dcell.fill = INPUT_FILL
    # auto roll-ups from Post Log, bounded to the 7-day window starting at B{row}
    win = (f"{PL}!$A${POSTLOG_FIRST}:$A${POSTLOG_LAST},\">=\"&$B{row},"
           f"{PL}!$A${POSTLOG_FIRST}:$A${POSTLOG_LAST},\"<\"&($B{row}+7)")
    wi.cell(row=row, column=3, value=f'=SUMIFS({PL}!$F${POSTLOG_FIRST}:$F${POSTLOG_LAST},{win})')
    wi.cell(row=row, column=4, value=(
        f'=SUMIFS({PL}!$G${POSTLOG_FIRST}:$G${POSTLOG_LAST},{win})'
        f'+SUMIFS({PL}!$H${POSTLOG_FIRST}:$H${POSTLOG_LAST},{win})'
        f'+SUMIFS({PL}!$I${POSTLOG_FIRST}:$I${POSTLOG_LAST},{win})'
        f'+SUMIFS({PL}!$J${POSTLOG_FIRST}:$J${POSTLOG_LAST},{win})'))
    wi.cell(row=row, column=5, value=f'=SUMIFS({PL}!$K${POSTLOG_FIRST}:$K${POSTLOG_LAST},{win})')
    # manual PostHog numbers
    for col in (6, 7):
        c = wi.cell(row=row, column=col)
        c.fill = INPUT_FILL
        c.border = BORDER
    # cumulative + ratios (auto)
    wi.cell(row=row, column=8,
            value=f'=IF($G{row}="","",{BASE}+SUM($G${WEEK_FIRST}:$G{row}))')
    wi.cell(row=row, column=9, value=f'=IFERROR($D{row}/$C{row},"")').number_format = PCT
    wi.cell(row=row, column=10, value=f'=IFERROR($G{row}/$F{row},"")').number_format = PCT
    for col in range(3, 11):
        cell = wi.cell(row=row, column=col)
        cell.border = BORDER
        if col in (3, 4, 5, 6, 7, 8):
            cell.number_format = NUM
        if col not in (6, 7):
            cell.fill = OUTPUT_FILL
# highlight signups-per-week below target pace handled on Path sheet
wi.freeze_panes = f"A{WEEK_FIRST}"
widths(wi, [8, 14, 13, 13, 16, 16, 15, 17, 14, 16])

WI = "'Weekly Inputs'"
WI_SIGN = f"{WI}!$G${WEEK_FIRST}:$G${WEEK_LAST}"
WI_CUM = f"{WI}!$H${WEEK_FIRST}:$H${WEEK_LAST}"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Funnel Dashboard
# ══════════════════════════════════════════════════════════════════════════════
fd = wb.create_sheet("Funnel Dashboard")
title(fd, "Funnel Dashboard", "All-time funnel from your data. Read top-to-bottom: where do people drop off?")
fd_headers = ["Funnel stage", "Total", "Conversion to this stage", "Target", "Status"]
HROW = 4
for i, h in enumerate(fd_headers, start=1):
    fd.cell(row=HROW, column=i, value=h)
style_header(fd, HROW, len(fd_headers))

C_SUM = f"SUM({WI}!$C${WEEK_FIRST}:$C${WEEK_LAST})"   # impressions
D_SUM = f"SUM({WI}!$D${WEEK_FIRST}:$D${WEEK_LAST})"   # engagements
E_SUM = f"SUM({WI}!$E${WEEK_FIRST}:$E${WEEK_LAST})"   # link clicks
F_SUM = f"SUM({WI}!$F${WEEK_FIRST}:$F${WEEK_LAST})"   # visitors
G_SUM = f"SUM({WI}!$G${WEEK_FIRST}:$G${WEEK_LAST})"   # signups

# stage rows: label, value-formula, conv-formula, target-ref, status-formula
rows = [
    ("Impressions", f"={C_SUM}", "", "", ""),
    ("Engagements (likes+comments+shares+saves)", f"={D_SUM}",
     f"=IFERROR({D_SUM}/{C_SUM},\"\")", f"={T_ENG}",
     f"=IF({D_SUM}=0,\"\",IF(IFERROR({D_SUM}/{C_SUM},0)>={T_ENG},\"On target\",\"Below\"))"),
    ("Link clicks", f"={E_SUM}",
     f"=IFERROR({E_SUM}/{C_SUM},\"\")", f"={T_CTR}",
     f"=IF({E_SUM}=0,\"\",IF(IFERROR({E_SUM}/{C_SUM},0)>={T_CTR},\"On target\",\"Below\"))"),
    ("Qualified visitors (PostHog)", f"={F_SUM}", "", "", ""),
    ("Waitlist signups  *** NORTH STAR ***", f"={BASE}+{G_SUM}",
     f"=IFERROR({G_SUM}/{F_SUM},\"\")", f"={T_CONV}",
     f"=IF({F_SUM}=0,\"\",IF(IFERROR({G_SUM}/{F_SUM},0)>={T_CONV},\"On target\",\"Below\"))"),
]
rr = HROW + 1
for label, val, conv, tgt, status in rows:
    fd.cell(row=rr, column=1, value=label).font = BOLD
    fd.cell(row=rr, column=2, value=val).number_format = NUM
    if conv:
        fd.cell(row=rr, column=3, value=conv).number_format = PCT
    if tgt:
        fd.cell(row=rr, column=4, value=tgt).number_format = PCT
    if status:
        fd.cell(row=rr, column=5, value=status)
    for col in range(1, 6):
        cell = fd.cell(row=rr, column=col)
        cell.border = BORDER
        if col in (2, 3, 4, 5):
            cell.fill = OUTPUT_FILL
    rr += 1
# colour the North Star row value
fd.cell(row=HROW + 5, column=2).font = Font(bold=True, size=12, color=BRAND)
# status colours
status_range = f"E{HROW+1}:E{rr-1}"
fd.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"On target"'], fill=GREEN_FILL, font=GREEN_FONT))
fd.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Below"'], fill=RED_FILL, font=RED_FONT))
fd.cell(row=rr + 1, column=1,
        value="Note: 'Waitlist signups' total includes the baseline already captured. "
              "Phase 2 will add Beta link clicks -> TestFlight installs -> Activated users below this line.").font = SUB_FONT
widths(fd, [40, 14, 22, 12, 12])

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Path to 2,000  (the headline)
# ══════════════════════════════════════════════════════════════════════════════
pt = wb.create_sheet("Path to 2,000")
title(pt, "Path to 2,000", "The one view for the founder: are we on pace to hit 2,000 signups before launch?")

def kv(row, label, formula, fmt=NUM, big=False):
    pt.cell(row=row, column=1, value=label).font = BOLD if not big else SECTION_FONT
    c = pt.cell(row=row, column=2, value=formula)
    c.fill = OUTPUT_FILL
    c.border = BORDER
    c.number_format = fmt
    if big:
        c.font = Font(bold=True, size=12, color=BRAND)
    return c

kv(4, "Goal — total signups", f"={GOAL}")
kv(5, "Signups captured so far", f"={BASE}+{G_SUM}", big=True)
kv(6, "Remaining to goal", f"={GOAL}-({BASE}+{G_SUM})")
kv(7, "Target launch date", f"={LAUNCH}", fmt=DATEFMT)
kv(8, "Weeks remaining to launch", f"=MAX(0,({LAUNCH}-TODAY())/7)", fmt="0.0")
kv(9, "Required signups / week (to hit goal)",
   f"=IF($B$8>0,$B$6/$B$8,$B$6)", big=True)
kv(10, "Current run-rate (avg signups / week)",
   f'=IFERROR(AVERAGE({WI_SIGN}),0)', big=True)
kv(11, "Projected weeks to reach goal (at run-rate)",
   f'=IFERROR($B$6/$B$10,"")', fmt="0.0")
kv(12, "Projected goal date (at run-rate)",
   f'=IF($B$10>0,TODAY()+$B$11*7,"")', fmt=DATEFMT)

pt.cell(row=14, column=1, value="STATUS").font = SECTION_FONT
st = pt.cell(row=14, column=2,
             value='=IF($B$10>=$B$9,"ON PACE",IF($B$10>=$B$9*0.8,"SLIGHTLY BEHIND","BEHIND"))')
st.border = BORDER
st.font = Font(bold=True, size=14)
st.alignment = Alignment(horizontal="center")
pt.conditional_formatting.add("B14",
    CellIsRule(operator="equal", formula=['"ON PACE"'], fill=GREEN_FILL, font=GREEN_FONT))
pt.conditional_formatting.add("B14",
    CellIsRule(operator="equal", formula=['"SLIGHTLY BEHIND"'], fill=AMBER_FILL, font=AMBER_FONT))
pt.conditional_formatting.add("B14",
    CellIsRule(operator="equal", formula=['"BEHIND"'], fill=RED_FILL, font=RED_FONT))

pt.cell(row=16, column=1,
        value="Run-rate = average of the weekly-signups you've entered so far. "
              "Enter at least 1 week before this means anything.").font = SUB_FONT

# cumulative-vs-goal line chart
chart = LineChart()
chart.title = "Cumulative signups by week"
chart.height = 8
chart.width = 18
data = Reference(wi, min_col=8, min_row=HROW, max_row=WEEK_LAST)   # col H + header row 4
cats = Reference(wi, min_col=2, min_row=WEEK_FIRST, max_row=WEEK_LAST)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.y_axis.title = "Signups"
pt.add_chart(chart, "D4")
widths(pt, [38, 16])

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — Channel Scorecard
# ══════════════════════════════════════════════════════════════════════════════
cs = wb.create_sheet("Channel Scorecard")
title(cs, "Channel Scorecard",
      "Which platform is actually working? Engagement auto-fills from Post Log; paste signups-by-source from PostHog (yellow).")
cs_headers = ["Platform", "Posts", "Impressions", "Engagements", "Engagement rate",
              "Link clicks", "Signups (from PostHog)", "Signups / 1k impressions"]
HROW = 4
for i, h in enumerate(cs_headers, start=1):
    cs.cell(row=HROW, column=i, value=h)
style_header(cs, HROW, len(cs_headers))

plat_win = (f"{PL}!$B${POSTLOG_FIRST}:$B${POSTLOG_LAST},$A{{r}}")
rr = HROW + 1
for plat in PLATFORMS:
    row = rr
    cs.cell(row=row, column=1, value=plat).font = BOLD
    crit = f'{PL}!$B${POSTLOG_FIRST}:$B${POSTLOG_LAST},$A{row}'
    cs.cell(row=row, column=2, value=f'=COUNTIFS({crit})')
    cs.cell(row=row, column=3, value=f'=SUMIFS({PL}!$F${POSTLOG_FIRST}:$F${POSTLOG_LAST},{crit})')
    cs.cell(row=row, column=4, value=(
        f'=SUMIFS({PL}!$G${POSTLOG_FIRST}:$G${POSTLOG_LAST},{crit})'
        f'+SUMIFS({PL}!$H${POSTLOG_FIRST}:$H${POSTLOG_LAST},{crit})'
        f'+SUMIFS({PL}!$I${POSTLOG_FIRST}:$I${POSTLOG_LAST},{crit})'
        f'+SUMIFS({PL}!$J${POSTLOG_FIRST}:$J${POSTLOG_LAST},{crit})'))
    cs.cell(row=row, column=5, value=f'=IFERROR($D{row}/$C{row},"")').number_format = PCT
    cs.cell(row=row, column=6, value=f'=SUMIFS({PL}!$K${POSTLOG_FIRST}:$K${POSTLOG_LAST},{crit})')
    sign = cs.cell(row=row, column=7)          # manual
    sign.fill = INPUT_FILL
    cs.cell(row=row, column=8, value=f'=IFERROR($G{row}/$C{row}*1000,"")').number_format = "0.00"
    for col in range(1, 9):
        cell = cs.cell(row=row, column=col)
        cell.border = BORDER
        if col in (2, 3, 4, 6):
            cell.number_format = NUM
        if col not in (1, 7):
            cell.fill = OUTPUT_FILL
    rr += 1
# Totals row
cs.cell(row=rr, column=1, value="TOTAL").font = BOLD
for col, letter in [(2, "B"), (3, "C"), (4, "D"), (6, "F"), (7, "G")]:
    cs.cell(row=rr, column=col,
            value=f'=SUM({letter}{HROW+1}:{letter}{rr-1})').number_format = NUM
cs.cell(row=rr, column=5, value=f'=IFERROR(D{rr}/C{rr},"")').number_format = PCT
cs.cell(row=rr, column=8, value=f'=IFERROR(G{rr}/C{rr}*1000,"")').number_format = "0.00"
for col in range(1, 9):
    cell = cs.cell(row=rr, column=col)
    cell.border = BORDER
    if col == 1:
        cell.font = BOLD
    else:
        cell.fill = OUTPUT_FILL
cs.cell(row=rr + 2, column=1,
        value="Tip: PostHog shows Instagram as 'ig'. The 'Signups (from PostHog)' column is the only "
              "thing you paste here weekly — everything else comes from Post Log.").font = SUB_FONT
widths(cs, [14, 8, 12, 13, 15, 11, 20, 22])

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7 — UTM Links  (extends taxonomy: adds Threads)
# ══════════════════════════════════════════════════════════════════════════════
ul = wb.create_sheet("UTM Links")
title(ul, "UTM Links",
      "Share THESE links (not plain attira.org) so PostHog can attribute every visitor & signup. Campaign: waitlist_launch")
BASE_URL = "https://attira.org/"
CAMPAIGN = "waitlist_launch"

def utm(source, medium, content=None):
    u = f"{BASE_URL}?utm_source={source}&utm_medium={medium}&utm_campaign={CAMPAIGN}"
    if content:
        u += f"&utm_content={content}"
    return u

LINKS = [
    ("LinkedIn", "linkedin", "social", ""),
    ("Twitter / X", "twitter", "social", ""),
    ("Threads", "threads", "social", ""),          # NEW
    ("Instagram", "instagram", "social", ""),
    ("Instagram (bio link)", "instagram", "bio", ""),
    ("Product Hunt", "producthunt", "referral", ""),
    ("Facebook", "facebook", "social", ""),
    ("Reddit (general)", "reddit", "social", ""),
]
SUBREDDITS = [
    ("r/femalefashionadvice", "femalefashionadvice"), ("r/fashion", "fashion"),
    ("r/OUTFITS", "outfits"), ("r/OOTD", "ootd"), ("r/whatshouldIwear", "whatshouldiwear"),
    ("r/capsulewardrobe", "capsulewardrobe"), ("r/minimalism", "minimalism"),
    ("r/declutter", "declutter"), ("r/Anticonsumption", "anticonsumption"),
    ("r/SideProject", "sideproject"), ("r/Startups", "startups"), ("r/artificial", "artificial"),
]
for label, content in SUBREDDITS:
    LINKS.append((label, "reddit", "social", content))

ul_headers = ["Channel", "utm_source", "utm_medium", "utm_campaign", "utm_content", "Full URL"]
HROW = 4
for i, h in enumerate(ul_headers, start=1):
    ul.cell(row=HROW, column=i, value=h)
style_header(ul, HROW, len(ul_headers))
rr = HROW + 1
for label, source, medium, content in LINKS:
    ul.cell(row=rr, column=1, value=label)
    ul.cell(row=rr, column=2, value=source)
    ul.cell(row=rr, column=3, value=medium)
    ul.cell(row=rr, column=4, value=CAMPAIGN)
    ul.cell(row=rr, column=5, value=content)
    link = utm(source, medium, content or None)
    c = ul.cell(row=rr, column=6, value=link)
    c.hyperlink = link
    c.font = Font(color=BRAND, underline="single")
    for col in range(1, 7):
        ul.cell(row=rr, column=col).border = BORDER
    rr += 1
ul.freeze_panes = f"A{HROW+1}"
widths(ul, [24, 14, 12, 16, 22, 80])

# ── save ───────────────────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "Attira_Growth_Dashboard.xlsx")
wb.save(out)
print("Wrote", out)

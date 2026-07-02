#!/usr/bin/env python3
"""
Build Attira_UTM_Tracking.xlsx — two tabs:
  1. UTM Links     — full share-link catalog (channels + Reddit subreddits)
  2. Performance   — live PostHog numbers (visitors / signups / conversion by source)

Link data is derived from website/utm-links.md.
Performance figures are pulled live from PostHog (project 454088) and passed in below;
re-run with refreshed numbers to update. Pass the "data as of" date as argv[1] (YYYY-MM-DD).
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_AS_OF = sys.argv[1] if len(sys.argv) > 1 else "unknown"

BASE = "https://attira.org/"
CAMPAIGN = "waitlist_launch"

def url(source, medium, content=None):
    u = "{}?utm_source={}&utm_medium={}&utm_campaign={}".format(BASE, source, medium, CAMPAIGN)
    if content:
        u += "&utm_content={}".format(content)
    return u

# (channel label, utm_source, utm_medium, utm_content)
LINKS = [
    ("LinkedIn",          "linkedin",    "social",   ""),
    ("Twitter / X",       "twitter",     "social",   ""),
    ("Instagram",         "instagram",   "social",   ""),
    ("Product Hunt",      "producthunt", "referral", ""),
    ("Facebook",          "facebook",    "social",   ""),
    ("Reddit (general)",  "reddit",      "social",   ""),
]
SUBREDDITS = [
    ("r/femalefashionadvice", "femalefashionadvice"),
    ("r/fashion",             "fashion"),
    ("r/OUTFITS",             "outfits"),
    ("r/OOTD",                "ootd"),
    ("r/whatshouldIwear",     "whatshouldiwear"),
    ("r/capsulewardrobe",     "capsulewardrobe"),
    ("r/minimalism",          "minimalism"),
    ("r/declutter",           "declutter"),
    ("r/Anticonsumption",     "anticonsumption"),
    ("r/SideProject",         "sideproject"),
    ("r/Startups",            "startups"),
    ("r/artificial",          "artificial"),
]
for label, content in SUBREDDITS:
    LINKS.append((label, "reddit", "social", content))

# ── Live PostHog data (last 90d, project 454088) ──────────────────────────────
# source -> (visitors, pageviews, signups)
PERF = [
    # source,          visitors, pageviews, signups
    ("ig",                  21,      34,      1),
    ("(direct / none)",     19,      29,      6),
    ("linkedin",             2,       2,      1),
    ("facebook",             1,       2,      0),
]
# every reddit subreddit currently has 0 traffic (links not shared yet)
TOTAL_UNIQUE_VISITORS = 41   # overall distinct persons (less than sum: people span sources)
TOTAL_EMAILS = 8

# ── Styling helpers ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1D4AFF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1D4AFF")
SUB_FONT = Font(italic=True, size=9, color="666666")
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER

def autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ══ Tab 1: UTM Links ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "UTM Links"
ws1["A1"] = "Attira — UTM Tracking Links"
ws1["A1"].font = TITLE_FONT
ws1["A2"] = "Campaign: {}  •  share these instead of the plain attira.org URL".format(CAMPAIGN)
ws1["A2"].font = SUB_FONT

headers1 = ["Channel", "utm_source", "utm_medium", "utm_campaign", "utm_content", "Full URL"]
hrow = 4
for i, h in enumerate(headers1, start=1):
    ws1.cell(row=hrow, column=i, value=h)
style_header(ws1, hrow, len(headers1))

r = hrow + 1
for label, source, medium, content in LINKS:
    ws1.cell(row=r, column=1, value=label)
    ws1.cell(row=r, column=2, value=source)
    ws1.cell(row=r, column=3, value=medium)
    ws1.cell(row=r, column=4, value=CAMPAIGN)
    ws1.cell(row=r, column=5, value=content)
    link = url(source, medium, content if content else None)
    cell = ws1.cell(row=r, column=6, value=link)
    cell.hyperlink = link
    cell.font = Font(color="1D4AFF", underline="single")
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=r, column=c).border = BORDER
    r += 1

ws1.freeze_panes = "A{}".format(hrow + 1)
autowidth(ws1, [22, 14, 12, 16, 20, 78])

# ══ Tab 2: Performance ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Performance")
ws2["A1"] = "Attira — Channel Performance"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "Live from PostHog (project 454088) • last 90 days • data as of {}".format(DATA_AS_OF)
ws2["A2"].font = SUB_FONT

# Summary block
ws2["A4"] = "Summary"
ws2["A4"].font = Font(bold=True, size=12)
summary = [
    ("Unique visitors (all sources)", TOTAL_UNIQUE_VISITORS),
    ("Waitlist emails captured", TOTAL_EMAILS),
    ("Overall visitor → email conversion", TOTAL_EMAILS / TOTAL_UNIQUE_VISITORS),
]
sr = 5
for label, val in summary:
    ws2.cell(row=sr, column=1, value=label).font = Font(bold=True)
    c = ws2.cell(row=sr, column=2, value=val)
    if "conversion" in label:
        c.number_format = "0.0%"
    sr += 1

# Per-source table
headers2 = ["Source", "Visitors", "Pageviews", "Signups", "Conversion %"]
hrow2 = sr + 1
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=hrow2, column=i, value=h)
style_header(ws2, hrow2, len(headers2))

rr = hrow2 + 1
for source, visitors, pageviews, signups in PERF:
    ws2.cell(row=rr, column=1, value=source)
    ws2.cell(row=rr, column=2, value=visitors)
    ws2.cell(row=rr, column=3, value=pageviews)
    ws2.cell(row=rr, column=4, value=signups)
    conv = ws2.cell(row=rr, column=5, value=(signups / visitors if visitors else 0))
    conv.number_format = "0.0%"
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=rr, column=c).border = BORDER
    rr += 1

# Reddit-by-subreddit note
rr += 1
ws2.cell(row=rr, column=1,
         value="Reddit subreddits (utm_content): 0 visits so far — links not yet shared. "
               "Once shared, filter utm_source=reddit and break down by utm_content.").font = SUB_FONT
rr += 1
ws2.cell(row=rr, column=1,
         value="Note: Instagram traffic arrives as 'ig' (not 'instagram'). Per-source visitor "
               "counts can sum above the unique total because a person may span sources.").font = SUB_FONT

ws2.freeze_panes = "A{}".format(hrow2 + 1)
autowidth(ws2, [34, 12, 12, 10, 14])

out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Attira_UTM_Tracking.xlsx")
wb.save(out)
print("Wrote", out)
